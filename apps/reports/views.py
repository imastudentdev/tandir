import json
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum, Q
from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from apps.sales.models import Sale, SaleItem, DebtPayment
from apps.production.models import ProductStock
from apps.expenses.models import Expense
from apps.users.models import CustomUser
from apps.users.mixins import AdminRequiredMixin
from apps.production.models import ProductionRecord


class AdminDashboardView(LoginRequiredMixin, AdminRequiredMixin, TemplateView): # O'zingizning AdminRequiredMixin ni qo'shib qo'ying
    """Adminlar va Direktorlar uchun boshqaruv paneli (Moliyaviy hisobotlar va tahlillar)"""
    template_name = "dashboard/reports.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        sales_query = Sale.objects.all()
        
        total_cash_sales = sales_query.filter(payment_type='cash').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        total_debt_sales = sales_query.filter(payment_type='debt').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        total_sales = total_cash_sales + total_debt_sales
        
        today_sales_cash = sales_query.filter(created_at__date=today, payment_type='cash').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        today_sales_debt = sales_query.filter(created_at__date=today, payment_type='debt').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
        today_total_sales = today_sales_cash + today_sales_debt

        total_collected_debt = DebtPayment.objects.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
        today_collected_debt = DebtPayment.objects.filter(created_at__date=today).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')

        active_market_debts = sales_query.filter(payment_type='debt', debt_status='unpaid').aggregate(total=Sum('remaining_debt'))['total'] or Decimal('0.00')

        total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        today_expenses = Expense.objects.filter(date=today).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        real_cash_in_hand = total_cash_sales + total_collected_debt - total_expenses
        today_cash_in_hand = today_sales_cash + today_collected_debt - today_expenses

        net_profit = total_sales - total_expenses
        today_net_profit = today_total_sales - today_expenses

        low_main_stock = ProductStock.objects.filter(quantity__lt=50).select_related('product')
        try:
            from apps.sales.models import SellerStock
            low_seller_stock = SellerStock.objects.filter(quantity__lt=30).select_related('product', 'seller')
        except ImportError:
            low_seller_stock = []

        top_debts = Sale.objects.filter(payment_type='debt', debt_status='unpaid').select_related('seller').order_by('-remaining_debt')[:5]

        top_products = SaleItem.objects.values('product__name').annotate(
            total_qty=Sum('quantity'),
            total_sum=Sum('subtotal')
        ).order_by('-total_qty')[:5]

        bakers_report = []
        bakers = CustomUser.objects.filter(role='baker')
        for baker in bakers:
            today_records = ProductionRecord.objects.filter(baker=baker, created_at__date=today)
            today_baked = today_records.aggregate(total=Sum('actual_quantity'))['total'] or 0
            today_earned = sum(r.actual_quantity * r.product.worker_share for r in today_records)
            
            unpaid_records = ProductionRecord.objects.filter(baker=baker, is_salary_calculated=False)
            unpaid_salary = sum(r.actual_quantity * r.product.worker_share for r in unpaid_records)
            
            bakers_report.append({
                'baker': baker,
                'today_baked': today_baked,
                'today_earned': today_earned,
                'unpaid_salary': unpaid_salary,
            })

        context.update({
            'total_cash': total_cash_sales,
            'total_collected_debt': total_collected_debt,
            'active_market_debts': active_market_debts,
            'real_cash_in_hand': real_cash_in_hand,
            'total_sales': total_sales,
            'net_profit': net_profit,
            
            'today_total_sales': today_total_sales,
            'today_expenses': today_expenses,
            'today_cash_in_hand': today_cash_in_hand,
            'today_net_profit': today_net_profit,
            
            'low_main_stock': low_main_stock,
            'low_seller_stock': low_seller_stock,
            'top_debts': top_debts,
            'top_products': top_products,
            'bakers_report': bakers_report,
            'is_loss': net_profit < 0,
            'is_today_loss': today_net_profit < 0,
        })
        return context


class ExportSalesReportExcelView(LoginRequiredMixin, AdminRequiredMixin, View): 
    """Sotuvlar tarixini professional va batafsil (Item-by-Item) darajada Excelga eksport qilish"""
    
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        
        ws = wb.active
        ws.title = "Batafsil Sotuvlar"
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        font_body = Font(name='Segoe UI', size=11)
        font_title = Font(name='Segoe UI', size=15, bold=True, color='1E3A8A')
        
        fill_header = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        fill_total = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
        
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
        )

        ws.merge_cells('A1:I1')
        ws['A1'] = "NONVOYXONA ERP TIZIMI - TRANSPARENT SOTUVLAR HISOBOTI (MAHSULOT KESIMIDA)"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 40
        ws.append([])

        headers = [
            "Chek №", "Mijoz Ismi", "Mijoz Telefoni", "Sotuvchi (Kassir)", 
            "Sotilgan Mahsulot", "Soni (dona)", "Asl Narxi (so'm)", "Oraliq Summa (so'm)", "Sotilgan Vaqt"
        ]
        ws.append(headers)
        ws.row_dimensions[3].height = 25
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        items = SaleItem.objects.select_related('sale', 'sale__seller', 'product').order_by('-sale_id')
        
        row_num = 4
        grand_total_qty = 0
        grand_total_sum = 0
        
        for item in items:
            sale = item.sale
            created_time = timezone.localtime(sale.created_at).strftime("%Y-%m-%d %H:%M")
            
            row_data = [
                f"#{sale.id}",
                sale.customer_name,
                sale.customer_phone,
                sale.seller.get_full_name() or sale.seller.username,
                item.product.name,
                item.quantity,
                item.unit_price,
                item.subtotal,
                created_time
            ]
            ws.append(row_data)
            
            grand_total_qty += item.quantity
            grand_total_sum += item.subtotal
            
            ws.cell(row=row_num, column=1).alignment = align_center
            ws.cell(row=row_num, column=2).alignment = align_left
            ws.cell(row=row_num, column=3).alignment = align_center
            ws.cell(row=row_num, column=4).alignment = align_left
            ws.cell(row=row_num, column=5).alignment = align_left
            
            qty_cell = ws.cell(row=row_num, column=6)
            qty_cell.alignment = align_center
            qty_cell.number_format = '#,##0'
            
            price_cell = ws.cell(row=row_num, column=7)
            price_cell.alignment = align_right
            price_cell.number_format = '#,##0.00'
            
            sub_cell = ws.cell(row=row_num, column=8)
            sub_cell.alignment = align_right
            sub_cell.number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=9).alignment = align_center
            
            for col_num in range(1, 10):
                c = ws.cell(row=row_num, column=col_num)
                c.font = font_body
                c.border = thin_border
            
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        ws.append(["UMUMIY JAMI", "", "", "", "", grand_total_qty, "", grand_total_sum, ""])
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        
        total_label = ws.cell(row=row_num, column=1)
        total_label.font = Font(name='Segoe UI', size=11, bold=True)
        total_label.alignment = align_left
        
        total_qty_cell = ws.cell(row=row_num, column=6)
        total_qty_cell.font = Font(name='Segoe UI', size=11, bold=True)
        total_qty_cell.alignment = align_center
        total_qty_cell.number_format = '#,##0'
        
        total_sum_cell = ws.cell(row=row_num, column=8)
        total_sum_cell.font = Font(name='Segoe UI', size=11, bold=True)
        total_sum_cell.alignment = align_right
        total_sum_cell.number_format = '#,##0.00'
        
        for col_num in range(1, 10):
            c = ws.cell(row=row_num, column=col_num)
            c.fill = fill_total
            c.border = thin_border
        ws.row_dimensions[row_num].height = 25

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column) 
            for cell in col:
                if cell.row == 1: continue
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        ws2 = wb.create_sheet(title="Nonvoylar Hisoboti")
        ws2.views.sheetView[0].showGridLines = True
        fill_baker_header = PatternFill(start_color='065F46', end_color='065F46', fill_type='solid')
        
        ws2.merge_cells('A1:E1')
        ws2['A1'] = "NONVOYLAR KUNLIK ISH HAQI VA QARZDORLIK BALANSLARI"
        ws2['A1'].font = font_title
        ws2['A1'].alignment = align_center
        ws2.row_dimensions[1].height = 40
        ws2.append([])

        baker_headers = ["Nonvoy Ismi", "Telefon Raqami", "Bugun Yopgan Noni (ta)", "Bugungi Ish Haqi (so'm)", "Jami Yopilmagan Haq (so'm)"]
        ws2.append(baker_headers)
        ws2.row_dimensions[3].height = 25
        
        for col_num, header in enumerate(baker_headers, 1):
            cell = ws2.cell(row=3, column=col_num)
            cell.font = font_header
            cell.fill = fill_baker_header
            cell.alignment = align_center
            cell.border = thin_border

        bakers = CustomUser.objects.filter(role='baker')
        b_row = 4
        today = timezone.now().date()
        
        for baker in bakers:
            today_records = ProductionRecord.objects.filter(baker=baker, created_at__date=today)
            t_baked = today_records.aggregate(total=Sum('actual_quantity'))['total'] or 0
            t_earned = sum(r.actual_quantity * r.product.worker_share for r in today_records)
            
            unpaid_records = ProductionRecord.objects.filter(baker=baker, is_salary_calculated=False)
            unpaid_s = sum(r.actual_quantity * r.product.worker_share for r in unpaid_records)
            
            ws2.append([
                baker.get_full_name() or baker.username,
                baker.phone_number if hasattr(baker, 'phone_number') else "-",
                t_baked,
                t_earned,
                unpaid_s
            ])
            
            ws2.cell(row=b_row, column=1).alignment = align_left
            ws2.cell(row=b_row, column=2).alignment = align_center
            ws2.cell(row=b_row, column=3).alignment = align_center
            
            c4 = ws2.cell(row=b_row, column=4)
            c4.alignment = align_right
            c4.number_format = '#,##0.00'
            
            c5 = ws2.cell(row=b_row, column=5)
            c5.alignment = align_right
            c5.number_format = '#,##0.00'
            
            for col_num in range(1, 6):
                c = ws2.cell(row=b_row, column=col_num)
                c.font = font_body
                c.border = thin_border
            b_row += 1

        for col in ws2.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1: continue
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 18)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Batafsil_ERP_Hisobot_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        
        return response